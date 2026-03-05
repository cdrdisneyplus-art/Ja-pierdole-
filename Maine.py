from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.checkbox import CheckBox
from kivy.uix.progressbar import ProgressBar
from kivy.clock import Clock

import os


class MainLayout(BoxLayout):

    def __init__(self, **kwargs):
        super().__init__(orientation="vertical", spacing=10, padding=10, **kwargs)

        self.selected_file = None
        self.headers = []
        self.rows = []
        self.checkboxes = []

        self.status = Label(text="Wybierz plik XLS lub XLSX")
        self.add_widget(self.status)

        btn_pick = Button(text="Wybierz plik")
        btn_pick.bind(on_press=self.pick_file)
        self.add_widget(btn_pick)

        btn_load = Button(text="Wgraj plik")
        btn_load.bind(on_press=self.load_file)
        self.add_widget(btn_load)

        self.scroll = ScrollView(size_hint=(1, 0.4))
        self.columns_layout = GridLayout(cols=2, size_hint_y=None)
        self.columns_layout.bind(minimum_height=self.columns_layout.setter("height"))

        self.scroll.add_widget(self.columns_layout)
        self.add_widget(self.scroll)

        self.progress = ProgressBar(max=100)
        self.add_widget(self.progress)

        btn_export = Button(text="Eksport")
        btn_export.bind(on_press=self.export_rows)
        self.add_widget(btn_export)

    def pick_file(self, instance):
        from plyer import filechooser

        filechooser.open_file(
            filters=["*.xls", "*.xlsx"],
            on_selection=self.handle_selection
        )

    def handle_selection(self, selection):
        if selection:
            self.selected_file = selection[0]
            name = os.path.basename(self.selected_file)
            self.status.text = f"Wybrano: {name}"

    def load_file(self, instance):

        if not self.selected_file:
            self.status.text = "Najpierw wybierz plik"
            return

        ext = self.selected_file.lower().split(".")[-1]

        self.headers = []
        self.rows = []

        try:

            if ext == "xlsx":
                from openpyxl import load_workbook

                wb = load_workbook(self.selected_file)
                sheet = wb.active

                self.headers = [cell.value for cell in sheet[1]]

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    self.rows.append(row)

            elif ext == "xls":
                import xlrd

                book = xlrd.open_workbook(self.selected_file)
                sheet = book.sheet_by_index(0)

                self.headers = sheet.row_values(0)

                for r in range(1, sheet.nrows):
                    self.rows.append(sheet.row_values(r))

            else:
                self.status.text = "Nieobsługiwany format"
                return

        except Exception as e:
            self.status.text = f"Błąd wczytywania: {e}"
            return

        self.build_column_selector()
        self.status.text = f"Wczytano rekordy: {len(self.rows)}"

    def build_column_selector(self):

        self.columns_layout.clear_widgets()
        self.checkboxes = []

        for header in self.headers:

            label = Label(text=str(header))
            checkbox = CheckBox(active=True)

            self.checkboxes.append(checkbox)

            self.columns_layout.add_widget(label)
            self.columns_layout.add_widget(checkbox)

    def export_rows(self, instance):

        if not self.rows:
            self.status.text = "Brak danych do eksportu"
            return

        from openpyxl import Workbook

        export_dir = "/storage/emulated/0/Documents/PaskiFuture"

        os.makedirs(export_dir, exist_ok=True)

        selected_columns = [
            i for i, cb in enumerate(self.checkboxes) if cb.active
        ]

        headers = [self.headers[i] for i in selected_columns]

        total = len(self.rows)

        for index, row in enumerate(self.rows):

            data = [row[i] for i in selected_columns]

            filename = "_".join(str(data[0]).split())

            if len(data) > 1:
                filename += "_" + "_".join(str(data[1]).split())

            filename += ".xlsx"

            path = os.path.join(export_dir, filename)

            wb = Workbook()
            ws = wb.active

            ws.append(headers)
            ws.append(data)

            wb.save(path)

            self.progress.value = int((index / total) * 100)

        self.progress.value = 100
        self.status.text = f"Eksport zakończony\n{export_dir}"


class PaskiFutureApp(App):

    def build(self):
        return MainLayout()

    def on_start(self):
        Clock.schedule_once(self.request_permissions, 1)

    def request_permissions(self, dt):

        try:
            from android.permissions import request_permissions, Permission

            request_permissions([
                Permission.READ_EXTERNAL_STORAGE,
                Permission.WRITE_EXTERNAL_STORAGE
            ])
        except:
            pass


if __name__ == "__main__":
    PaskiFutureApp().run()
